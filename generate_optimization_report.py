#!/usr/bin/env python3
"""
UniFi Network Optimization Report Generator
Fetches live diagnostics and produces a ranked XLSX with highest-impact
channel changes first (20-80 rule: the few APs causing most congestion).
"""

import json, urllib.request, ssl, sys, os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Fetch live data ──────────────────────────────────────────────────────────
API = "http://localhost:3000/api/diagnostics"
ctx = ssl.create_default_context(); ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
with urllib.request.urlopen(API, timeout=10) as r:
    data = json.loads(r.read())

radios   = data["channels"]["radios"]
summary  = data["channels"]["summary"]
clients  = data.get("clients", {}).get("clients", [])

# ── Impact scoring ───────────────────────────────────────────────────────────
# Score = cu_total * (1 + tx_retries_pct/100) * cci_weight
# cci_weight: more co-channel interferers → higher impact of fixing this AP
# 5GHz weights reflect non-DFS channels only (36-48 UNII-1, 149-165 UNII-3)
CCI_WEIGHT = {"ng": {6: 13, 11: 9, 1: 11}, "na": {44: 17, 40: 9, 36: 1, 48: 3, 149: 3, 153: 3, 157: 3, 161: 3, 165: 3}}

def impact_score(r):
    cu   = r.get("cu_total", 0)
    ret  = r.get("tx_retries_pct", 0) or 0
    cci  = r.get("cci_count", 1) or 1
    # Disabled radios → 0
    if not r.get("channel"):
        return 0
    return round(cu * (1 + ret / 100) * (cci / 5), 1)

for r in radios:
    r["_impact"] = impact_score(r)

# Sort by impact descending
radios_sorted = sorted([r for r in radios if r.get("channel")],
                       key=lambda x: x["_impact"], reverse=True)

# ── Channel recommendation logic ─────────────────────────────────────────────
# Non-DFS channels only — DFS channels (52–144) are excluded because
# iPads and many client devices cannot associate on DFS channels.
CHAN_24 = [1, 6, 11]
CHAN_5  = [36, 40, 44, 48, 149, 153, 157, 161, 165]

ch_usage_24 = summary["channelCounts24"]
ch_usage_5  = summary["channelCounts5"]

def best_channel(radio):
    if radio["band"] == "2.4GHz":
        return min(CHAN_24, key=lambda c: ch_usage_24.get(str(c), 0))
    else:
        return min(CHAN_5, key=lambda c: ch_usage_5.get(str(c), 0))

# ── Build AP → client map ────────────────────────────────────────────────────
ap_clients = {}
for c in clients:
    ap = c.get("apName", "")
    ap_clients.setdefault(ap, []).append(c)

def ap_struggling(ap_name):
    cc = ap_clients.get(ap_name, [])
    bad = [c for c in cc if c.get("severity") in ("critical", "warning")]
    return len(cc), len(bad)

# ── Colors ───────────────────────────────────────────────────────────────────
RED    = PatternFill("solid", fgColor="C0392B")
ORANGE = PatternFill("solid", fgColor="E67E22")
YELLOW = PatternFill("solid", fgColor="F1C40F")
GREEN  = PatternFill("solid", fgColor="27AE60")
BLUE   = PatternFill("solid", fgColor="2980B9")
GRAY   = PatternFill("solid", fgColor="BDC3C7")
HEADER = PatternFill("solid", fgColor="1A252F")
SUB_H  = PatternFill("solid", fgColor="2C3E50")

WHITE  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
DARK   = Font(color="1A252F", bold=True, name="Calibri", size=10)
NORMAL = Font(name="Calibri", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="7F8C8D")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def health_fill(h):
    return {"critical": RED, "warning": ORANGE, "healthy": GREEN}.get(h, GRAY)

def severity_fill(s):
    return {"critical": RED, "warning": ORANGE, "healthy": GREEN}.get(s, GRAY)

def pct_fill(v):
    if v >= 50: return RED
    if v >= 30: return ORANGE
    if v >= 15: return YELLOW
    return GREEN

# ── Create workbook ───────────────────────────────────────────────────────────
wb  = openpyxl.Workbook()
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Channel Optimization Blueprint (sorted by impact)
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📡 Channel Optimization"

ws1.merge_cells("A1:L1")
ws1["A1"] = f"Tailored Optimal Network Channel Grid Blueprint  •  {NOW}"
ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
ws1["A1"].fill = HEADER
ws1["A1"].alignment = CENTER
ws1.row_dimensions[1].height = 28

ws1.merge_cells("A2:L2")
ws1["A2"] = (
    "Ranked by IMPACT SCORE (most congested → most benefit if changed first). "
    "Changing the top 20% of APs eliminates ~80% of interference."
)
ws1["A2"].font = Font(name="Calibri", size=9, italic=True, color="FFFFFF")
ws1["A2"].fill = SUB_H
ws1["A2"].alignment = CENTER
ws1.row_dimensions[2].height = 16

cols = ["#", "AP Name", "IP", "Model", "Band", "Current Ch",
        "CU Total %", "TX Retry %", "Co-Ch Peers", "Impact Score",
        "Suggested Ch", "Health"]
widths = [4, 28, 14, 14, 7, 10, 11, 11, 12, 13, 12, 10]

for ci, (col, w) in enumerate(zip(cols, widths), 1):
    c = ws1.cell(row=3, column=ci, value=col)
    c.font = WHITE; c.fill = HEADER; c.alignment = CENTER; c.border = BORDER
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.row_dimensions[3].height = 22

rank = 0
for ri, r in enumerate(radios_sorted, 1):
    if r["_impact"] == 0:
        continue
    rank += 1
    row = ri + 3
    suggest = best_channel(r)
    ch = r.get("channel")
    same = (suggest == ch)

    vals = [
        rank, r["apName"], r["ip"], r["model"],
        r["band"], ch,
        r.get("cu_total", 0), round(r.get("tx_retries_pct", 0) or 0, 1),
        r.get("cci_count", 0),
        r["_impact"],
        ch if same else suggest,
        r.get("health", "—")
    ]
    fills = [
        None, None, None, None, None,
        pct_fill(r.get("cu_total", 0)),
        pct_fill(r.get("cu_total", 0)),
        pct_fill(r.get("tx_retries_pct", 0) or 0),
        ORANGE if (r.get("cci_count", 0) or 0) >= 10 else GREEN,
        RED if r["_impact"] >= 80 else (ORANGE if r["_impact"] >= 40 else YELLOW if r["_impact"] >= 15 else GRAY),
        GREEN if same else BLUE,
        health_fill(r.get("health", ""))
    ]

    for ci, (v, f) in enumerate(zip(vals, fills), 1):
        cell = ws1.cell(row=row, column=ci, value=v)
        cell.font = NORMAL; cell.alignment = CENTER; cell.border = BORDER
        if f: cell.fill = f
        if ci in (7, 8) and isinstance(v, (int, float)):
            cell.number_format = "0.0\"%\""
    ws1.row_dimensions[row].height = 18

ws1.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Client Connectivity Issues
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🔌 Client Issues")
ws2.merge_cells("A1:N1")
ws2["A1"] = f"Client Connectivity Report  •  {NOW}"
ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
ws2["A1"].fill = HEADER
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 28

cc = ["Hostname", "IP", "Severity", "AP Name", "Band", "Channel",
      "Signal (dBm)", "Satisfaction", "TX Retry %", "Roams",
      "RX Rate (Mbps)", "TX Rate (Mbps)", "Data Used (MB)", "Issues"]
cw = [18, 15, 10, 28, 7, 9, 13, 13, 11, 8, 14, 14, 15, 60]

for ci, (col, w) in enumerate(zip(cc, cw), 1):
    c = ws2.cell(row=2, column=ci, value=col)
    c.font = WHITE; c.fill = HEADER; c.alignment = CENTER; c.border = BORDER
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[2].height = 22

# Sort clients: critical first, then warning, then by roamCount desc
sev_order = {"critical": 0, "warning": 1, "healthy": 2}
clients_sorted = sorted(clients, key=lambda x: (
    sev_order.get(x.get("severity","healthy"), 2),
    -x.get("roamCount", 0)
))

for ri, c in enumerate(clients_sorted, 3):
    sev = c.get("severity", "healthy")
    sf  = severity_fill(sev)
    flags = "; ".join(c.get("flags", [])) or "—"
    vals = [
        c.get("hostname", "?"),
        c.get("ip", "—"),
        sev.upper(),
        c.get("apName", "—"),
        c.get("band", "—"),
        c.get("channel", "—"),
        c.get("signal", "—"),
        c.get("satisfaction", "—"),
        round(c.get("txRetriesPct", 0) or 0, 1),
        c.get("roamCount", 0),
        round((c.get("rxRateKbps", 0) or 0) / 1000, 1),
        round((c.get("txRateKbps", 0) or 0) / 1000, 1),
        round((c.get("totalBytes", 0) or 0) / 1_048_576, 1),
        flags
    ]
    for ci2, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci2, value=v)
        cell.font = NORMAL; cell.border = BORDER
        cell.alignment = LEFT if ci2 == 14 else CENTER
        if ci2 == 3:   cell.fill = sf; cell.font = WHITE
        elif ci2 == 8: cell.fill = pct_fill(100 - (v or 100))
        elif ci2 == 9: cell.fill = pct_fill(v or 0)
    ws2.row_dimensions[ri].height = 18

ws2.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Summary Dashboard
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📊 Summary")
ws3.merge_cells("A1:D1")
ws3["A1"] = f"Network Health Summary  •  {NOW}"
ws3["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
ws3["A1"].fill = HEADER; ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 28

kv = [
    ("Total APs",        summary.get("totalAPs")),
    ("Active 2.4GHz Radios", summary.get("totalRadios24")),
    ("Active 5GHz Radios",   summary.get("totalRadios5")),
    ("Avg Utilization 2.4GHz", f"{summary.get('avgUtil24',0)}%"),
    ("Avg Utilization 5GHz",   f"{summary.get('avgUtil5',0)}%"),
    ("Congested Radios",  summary.get("congestedRadiosCount")),
    ("Warning Radios",    summary.get("warningRadiosCount")),
    ("Total Clients",    data.get("clients",{}).get("summary",{}).get("totalAllClients","?")),
    ("Critical Clients", data.get("clients",{}).get("summary",{}).get("criticalCount","?")),
    ("Warning Clients",  data.get("clients",{}).get("summary",{}).get("warningCount","?")),
    ("Healthy Clients",  data.get("clients",{}).get("summary",{}).get("healthyCount","?")),
    ("Health Index",     f"{data.get('clients',{}).get('summary',{}).get('healthIndex','?')}%"),
    # Channel distribution 2.4GHz
    ("2.4GHz CH-1 APs",  ch_usage_24.get("1",0)),
    ("2.4GHz CH-6 APs",  ch_usage_24.get("6",0)),
    ("2.4GHz CH-11 APs", ch_usage_24.get("11",0)),
    # Channel distribution 5GHz (top used)
    ("5GHz CH-44 APs",  ch_usage_5.get("44",0)),
    ("5GHz CH-40 APs",  ch_usage_5.get("40",0)),
    ("5GHz CH-108 APs", ch_usage_5.get("108",0)),
    ("5GHz CH-60 APs",  ch_usage_5.get("60",0)),
]

for i, (k, v) in enumerate(kv, 2):
    ws3.cell(row=i, column=1, value=k).font = Font(name="Calibri", size=10, bold=True)
    ws3.cell(row=i, column=1).fill = SUB_H
    ws3.cell(row=i, column=1).font = WHITE
    ws3.cell(row=i, column=2, value=v).font = NORMAL
    ws3.cell(row=i, column=1).border = BORDER
    ws3.cell(row=i, column=2).border = BORDER
    ws3.row_dimensions[i].height = 18

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 16

# Save
out = os.path.join(os.path.dirname(__file__), "unifi_optimization_report.xlsx")
wb.save(out)
print(f"\n[OK] Report saved -> {out}")
print(f"   {rank} radios ranked by impact | {len(clients_sorted)} clients analyzed")
